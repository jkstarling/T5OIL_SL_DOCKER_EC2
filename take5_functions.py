# import pandas as pd
import numpy as np


def create_all_months_df(locations, start_month, end_month, PLBAL = 'PL'):  # use 'PL' or 'BAL'
    '''
    Creates a dataframe with all months represented to make sure 
    that the T5 P&L shows all months for locations that don't have records for every month. 
    '''
    # import pandas as pd
    import numpy as np
    import pandas as pd

    if PLBAL == 'PL':
        myacct = 'Income'
    elif PLBAL == 'BAL':
        myacct = 'ASSETS'
    else:
        print('Value for PLBAL must be "PL" or "BAL"')

    # make dataframe with all months and all locations to make sure the pivot will have all months present
    all_MM = pd.date_range(start=start_month, end=end_month, freq='MS').strftime('%b %Y')
    all_mm = pd.to_datetime(all_MM, format='%b %Y')
    mm_df = pd.DataFrame({'Month': all_MM, 'monthdt': all_mm,  'Account': myacct, 'value': np.nan})

    inc_df = pd.DataFrame()
    for dummysheet, dummysheet2, loc in locations:
        # print(loc)
        mm_df['location'] = str(loc)
        inc_df = pd.concat([inc_df, mm_df], ignore_index=True)
    return inc_df


def read_in_PL_files(PL_folder_path, inc_df, controlmap, ex_acct_map):
    '''
    Reads in all excel files that are found in the PL folder path.
    Does necessary data cleaning, including:
        removes "=" from QBO output.
        makes 0.0's NaN
        strips account numbers (4 #s) from the account column
        (and some other stuff)
    and returns the "result" dataframe. 
    '''
    import pandas as pd
    import numpy as np
    import os
    from openpyxl import load_workbook

    # Get a list of all Excel files in the directory
    myfiles = [file for file in os.listdir(PL_folder_path) if file.endswith('.xlsx')]

    # Create an empty list to store dataframes
    dfs = []

    # Loop through each Excel file and read it into a dataframe
    for file in myfiles:
        file_path = os.path.join(PL_folder_path, file)
        # get month + year
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        mymth = sheet.cell(row=3, column=1).value
        # Assign 'Account' to cell A4
        sheet.cell(row=5, column=1).value = 'Account'
        # Read in the rest of the data
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[4]
        df = df[5:]  
        df['Month'] = mymth
        df['monthdt'] = pd.to_datetime(mymth, format='%B %Y')
        df.drop(columns="Total", inplace=True)
        df_melt = pd.melt(df, id_vars=['Month', 'monthdt', 'Account'], var_name='location', value_name='value')
        # Remove "=" from the column 'column_name'
        df_melt['value'] = df_melt['value'].str.replace('=', '')
        dfs.append(df_melt)   # Append the dataframe to the list
        # Save the workbook with the changes
        # workbook.save(file_path)
        print(file)

    # Concatenate all dataframes in the list
    result = pd.concat(dfs, ignore_index=True)
    # Strip whitespace from string columns
    result = result.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    # drop all Income columns and replace with the "inc_df" entries made above
    result = result[result.Account != 'Income']  
    result = pd.concat([result, inc_df], ignore_index=True)
    result.reset_index(drop=True, inplace=True)

    result['Account_Num'] = pd.to_numeric(result['Account'].str.slice(0,4), errors='coerce')
    result.reset_index(drop=True, inplace=True)

    mask_non_numeric = ~pd.to_numeric(result['value'], errors='coerce').notna()
    # Replace non-numeric values with NaN
    result.loc[mask_non_numeric, 'value'] = np.nan
    # Check if 'value' column is None
    mask_none_value = result['value'].isna()
    # Check if 'Account' column has an entry in 'ex_acct_map'
    mask_account_in_map = result['Account'].isin(ex_acct_map['Account'])
    # Apply conditions and replace 'Account_Num' values accordingly
    result.loc[mask_account_in_map, 'Account_Num'] = result['Account'].map(ex_acct_map.set_index('Account')['Account_Num'])
    # result.loc[mask_none_value & mask_account_in_map, 'Account_Num'] = result['Account'].map(ex_acct_map.set_index('Account')['Account_Num'])

    result = result.dropna(subset=['Account_Num']).reset_index(drop=True) # drop NaNs for accounts (extra lines)
    result.replace(0.0, np.nan, inplace=True)  # change 0.0 to NaN
    result.value = pd.to_numeric(result.value).reset_index(drop=True)
    result.sort_values(by='monthdt', inplace=True)
    result = result.merge(controlmap, on='Account', how='left')
    # Display the updated DataFrame
    # print(result)
    return result


def read_in_BAL_files(PL_folder_path, inc_df, ex_acct_map):
    '''
    Reads in all excel files that are found in the BAL folder path.
    Does necessary data cleaning, including:
        removes "=" from QBO output.
        makes 0.0's NaN
        strips account numbers (4 #s) from the account column
        (and some other stuff)
    and returns the "result" dataframe. 
    '''
    import pandas as pd
    import numpy as np
    import os
    from openpyxl import load_workbook
    import re

    # Get a list of all Excel files in the directory
    myfiles = [file for file in os.listdir(PL_folder_path) if file.endswith('.xlsx')]

    # Create an empty list to store dataframes
    dfs = []

    # Loop through each Excel file and read it into a dataframe
    for file in myfiles:
        file_path = os.path.join(PL_folder_path, file)
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        # Assign 'Account' to cell A4
        sheet.cell(row=5, column=1).value = 'Account'
        # Read in the rest of the data
        df = pd.DataFrame(sheet.values)
        df.columns = df.iloc[4]
        df = df[5:]  
        # get month + year
        mymth = sheet.cell(row=3, column=1).value
        date_match = re.search(r'(\w+ \d{1,2}, \d{4})', mymth)
        # print(date_match)
        if date_match:
            extracted_date = pd.to_datetime(date_match.group(1), format='%B %d, %Y').strftime('%B %Y')
            # print(extracted_date)
            # Convert extracted date to datetime object
            df['monthdt'] = pd.to_datetime(extracted_date, format='%B %Y')
        else:
            print("Date not found in the string.")
        df['Month'] = extracted_date
        # df['monthdt'] = pd.to_datetime(mymth, format='%b %Y')
        df.drop(columns="Total", inplace=True)
        df_melt = pd.melt(df, id_vars=['Month', 'monthdt', 'Account'], var_name='location', value_name='value')
        # Remove "=" from the column 'column_name'
        df_melt['value'] = df_melt['value'].str.replace('=', '')
        dfs.append(df_melt)   # Append the dataframe to the list
        # Save the workbook with the changes
        # workbook.save(file_path)
        print(file)

    # Concatenate all dataframes in the list
    result = pd.concat(dfs, ignore_index=True)
    # Strip whitespace from string columns
    result = result.applymap(lambda x: x.strip() if isinstance(x, str) else x)

    # drop all Income columns and replace with the "inc_df" entries made above
    result = result[result.Account != 'ASSETS']  
    result = pd.concat([result, inc_df], ignore_index=True)
    result.reset_index(drop=True, inplace=True)

    # ind = (result.Account == 'Net Income')
    # print(result[ind])
    result['Account_Num'] = pd.to_numeric(result['Account'].str.slice(0,4), errors='coerce')
    result.reset_index(drop=True, inplace=True)

    mask_non_numeric = ~pd.to_numeric(result['value'], errors='coerce').notna()
    # Replace non-numeric values with NaN
    result.loc[mask_non_numeric, 'value'] = np.nan
    # Check if 'value' column is None
    mask_none_value = result['value'].isna()
    # Check if 'Account' column has an entry in 'ex_acct_map'
    mask_account_in_map = result['Account'].isin(ex_acct_map['Account'])
    # Apply conditions and replace 'Account_Num' values accordingly
    result.loc[mask_account_in_map, 'Account_Num'] = result['Account'].map(ex_acct_map.set_index('Account')['Account_Num'])
    # result.loc[mask_none_value & mask_account_in_map, 'Account_Num'] = result['Account'].map(ex_acct_map.set_index('Account')['Account_Num'])

    
    # ind = (result.Account == 'Net Income')
    # print(result[ind])

    result = result.dropna(subset=['Account_Num']).reset_index(drop=True) # drop NaNs for accounts (extra lines)
    result.replace(0.0, np.nan, inplace=True)  # change 0.0 to NaN
    result.value = pd.to_numeric(result.value).reset_index(drop=True)
    result.sort_values(by='monthdt', inplace=True)
    # result = result.merge(controlmap, on='Account', how='left')
    # Display the updated DataFrame
    # print(result)
    return result


def create_T5_PL_pivot_table(result_df, myloc, controlmap):
    '''
    Takes the result dataframe and the T5 location and returns 
    a pivot dataframe across all months
    '''
    # import pandas as pd
    import numpy as np

    result = result_df
    # Pivot the DataFrame by 'Month', keeping columns ('Account_Num', 'Account'), summing 'value', and filtering by 'location'
    pivot_table = result[result['location'] == myloc].pivot_table(index=['Account_Num', 'Account'], columns='monthdt', values='value', aggfunc='sum')

    ### change the dt to Mmmm YYYY 
    # pivot_table.replace(0.0, np.nan, inplace=True)
    pivot_table = pivot_table.reindex(columns=pivot_table.columns.sort_values())
    pivot_table.columns = pivot_table.columns.strftime('%b %y')

    ### 4999  Total Income	(sum all 4000s)
    # Filter the pivot table by "Account_Num" greater than 4000 and less than 4999
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 4000) & 
        (pivot_table.index.get_level_values('Account_Num') < 4999)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(4999, 'Total Income'), :] = summed_values

    ### 5998 Total Cost of Goods Sold	(sum all 5000s)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 5000) & 
        (pivot_table.index.get_level_values('Account_Num') < 5998)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(5998, 'Total Cost of Goods Sold'), :] = summed_values

    ###  5999  Gross Profit	5998  minus 4999
    top = pivot_table.loc[(4999, 'Total Income'), :]
    bot = pivot_table.loc[(5998, 'Total Cost of Goods Sold'), :]
    pivot_table.loc[(5999, 'Gross Profit'), :] = top - bot

    ### 8990  Total Expenses	(add all 6000 - 8980(inclusive))
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 6000) & 
        (pivot_table.index.get_level_values('Account_Num') < 8990)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(8990, 'Total Expenses'), :] = summed_values

    ### 8999  Net Operating Income	5999 minus 8990  
    top = pivot_table.loc[(5999, 'Gross Profit'), :]
    bot = pivot_table.loc[(8990, 'Total Expenses'), :]
    pivot_table.loc[(8999, 'Net Operating Income'), :] = top - bot

    ### 9199  Total Other Income	  (add all 9100s)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 9000) & 
        (pivot_table.index.get_level_values('Account_Num') < 9199)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(9199, 'Total Other Income'), :] = summed_values

    ### 9980  Total Other Expenses  (add all 9200s + )
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 9200) & 
        (pivot_table.index.get_level_values('Account_Num') < 9980)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(9980, 'Total Other Expenses'), :] = summed_values

    ###  9990    Net Other Income	 9199 minus  9980  
    top = pivot_table.loc[(9199, 'Total Other Income'), :]
    bot = pivot_table.loc[(9980, 'Total Other Expenses'), :]
    pivot_table.loc[(9990, 'Net Other Income'), :] = top - bot

    ###  9999   Net Income	           9990 plus 8999
    top = pivot_table.loc[(9990, 'Net Other Income'), :]
    bot = pivot_table.loc[(8999, 'Net Operating Income'), :]
    pivot_table.loc[(9999, 'Net Income'), :] = top + bot

    # remove all 0.0s and replace with NaNs
    pivot_table.replace(0.0, np.nan, inplace=True)


    ##### create 4 wall reports #####

    # make 4 blank rows
    for i in range(10000, 10004):
        pivot_table.loc[(i, ''), :] = [None] * len(pivot_table.columns)


    # 10005	4 Wall Expenses	(all 6000 to 7999)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 6000) & 
        (pivot_table.index.get_level_values('Account_Num') < 7999)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(10005, '4 Wall Expenses'), :] = summed_values

    # 10006	4 Wall Profit	Gross Profit 5999 minus 10005
    top = pivot_table.loc[(5999, 'Gross Profit'), :]
    bot = pivot_table.loc[(10005, '4 Wall Expenses'), :]
    pivot_table.loc[(10006, '4 Wall Profit'), :] = top - bot

    # 10011	Labor	(add all 6000s)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 6000) & 
        (pivot_table.index.get_level_values('Account_Num') < 6999)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(10011, 'Labor'), :] = summed_values

    # 10012	Controllable	(add all 7000s that are CONTROLABLE)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 7000) & 
        (pivot_table.index.get_level_values('Account_Num') < 7999)
    ]
    # add along columns if Controllable
    existing_accounts = filtered_pivot_table.index.get_level_values(1)
    controllable_rows = controlmap[(controlmap['Account'].isin(existing_accounts)) & (controlmap['Controllable'] == 'Controllable')]
    account_names = controllable_rows['Account'].tolist()
    account_names = [(int(x[:4]), x) for x in account_names]
    selected_rows = filtered_pivot_table.loc[account_names,:]
    sum_along_columns = selected_rows.sum()
    pivot_table.loc[(10012, 'Controllable'), :] = sum_along_columns

    # 10013	Uncontrollable	(add all 7000s that are UNCONTROLABLE)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 7000) & 
        (pivot_table.index.get_level_values('Account_Num') < 7999)
    ]
    # add along columns if Controllable
    existing_accounts = filtered_pivot_table.index.get_level_values(1)
    controllable_rows = controlmap[(controlmap['Account'].isin(existing_accounts)) & (controlmap['Controllable'] == 'Uncontrollable')]
    account_names = controllable_rows['Account'].tolist()
    account_names = [(int(x[:4]), x) for x in account_names]
    selected_rows = filtered_pivot_table.loc[account_names,:]
    sum_along_columns = selected_rows.sum()
    pivot_table.loc[(10013, 'Uncontrollable'), :] = sum_along_columns

    # 10021	Labor %	10011 divide by  4999  (4999, 'Total Income')
    top = pivot_table.loc[(10011, 'Labor'), :]
    bot = pivot_table.loc[(4999, 'Total Income'), :]
    pivot_table.loc[(10021, 'Labor %'), :] = top / bot

    # 10022	Controllable %	10012 divide by  4999  (4999, 'Total Income')
    top = pivot_table.loc[(10012, 'Controllable'), :]
    bot = pivot_table.loc[(4999, 'Total Income'), :]
    pivot_table.loc[(10022, 'Controllable %'), :] = top / bot

    # 10023	Uncontrollable %	10013 divide by  4999  (4999, 'Total Income')
    top = pivot_table.loc[(10013, 'Uncontrollable'), :]
    bot = pivot_table.loc[(4999, 'Total Income'), :]
    pivot_table.loc[(10023, 'Uncontrollable %'), :] = top / bot

    # Sort the index
    pivot_table.sort_index(inplace=True)

    return(pivot_table)


def create_T5_PL_ALL_pivot_table(result_df, controlmap):
    '''
    Takes the result dataframe and the T5 location and returns 
    a pivot dataframe across all months
    '''
    # import pandas as pd
    import numpy as np

    result = result_df
    # Pivot the DataFrame by 'Month', keeping columns ('Account_Num', 'Account'), summing 'value', and filtering by 'location'
    pivot_table = result.pivot_table(index=['Account_Num', 'Account'], columns='monthdt', values='value', aggfunc='sum')

    ### change the dt to Mmmm YYYY 
    # pivot_table.replace(0.0, np.nan, inplace=True)
    pivot_table = pivot_table.reindex(columns=pivot_table.columns.sort_values())
    pivot_table.columns = pivot_table.columns.strftime('%b %y')

    ### 4999  Total Income	(sum all 4000s)
    # Filter the pivot table by "Account_Num" greater than 4000 and less than 4999
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 4000) & 
        (pivot_table.index.get_level_values('Account_Num') < 4999)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(4999, 'Total Income'), :] = summed_values      ### KEEP ###

    ### 5998 Total Cost of Goods Sold	(sum all 5000s)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 5000) & 
        (pivot_table.index.get_level_values('Account_Num') < 5998)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(5998, 'Total Cost of Goods Sold'), :] = summed_values

    ###  5999  Gross Profit	5998  minus 4999
    top = pivot_table.loc[(4999, 'Total Income'), :]
    bot = pivot_table.loc[(5998, 'Total Cost of Goods Sold'), :]
    pivot_table.loc[(5999, 'Gross Profit'), :] = top - bot   ### KEEP ###

    ### 8990  Total Expenses	(add all 6000 - 8980(inclusive))
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 6000) & 
        (pivot_table.index.get_level_values('Account_Num') < 8990)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(8990, 'Total Expenses'), :] = summed_values

    ### 8999  Net Operating Income	5999 minus 8990  
    top = pivot_table.loc[(5999, 'Gross Profit'), :]
    bot = pivot_table.loc[(8990, 'Total Expenses'), :]
    pivot_table.loc[(8999, 'Net Operating Income'), :] = top - bot

    ### 9199  Total Other Income	  (add all 9100s)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 9000) & 
        (pivot_table.index.get_level_values('Account_Num') < 9199)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(9199, 'Total Other Income'), :] = summed_values

    ### 9980  Total Other Expenses  (add all 9200s + )
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 9200) & 
        (pivot_table.index.get_level_values('Account_Num') < 9980)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(9980, 'Total Other Expenses'), :] = summed_values

    ###  9990    Net Other Income	 9199 minus  9980  
    top = pivot_table.loc[(9199, 'Total Other Income'), :]
    bot = pivot_table.loc[(9980, 'Total Other Expenses'), :]
    pivot_table.loc[(9990, 'Net Other Income'), :] = top - bot

    ###  9999   Net Income	           9990 plus 8999
    top = pivot_table.loc[(9990, 'Net Other Income'), :]
    bot = pivot_table.loc[(8999, 'Net Operating Income'), :]
    pivot_table.loc[(9999, 'Net Income'), :] = top + bot

    # remove all 0.0s and replace with NaNs
    pivot_table.replace(0.0, np.nan, inplace=True)


    ##### create 4 wall reports #####

    # make 4 blank rows
    for i in range(10000, 10004):
        pivot_table.loc[(i, ''), :] = [None] * len(pivot_table.columns)


    # 10005	4 Wall Expenses	(all 6000 to 7999)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 6000) & 
        (pivot_table.index.get_level_values('Account_Num') < 7999)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(10005, '4 Wall Expenses'), :] = summed_values

    # 10006	4 Wall Profit	Gross Profit 5999 minus 10005
    top = pivot_table.loc[(5999, 'Gross Profit'), :]
    bot = pivot_table.loc[(10005, '4 Wall Expenses'), :]
    pivot_table.loc[(10006, '4 Wall Profit'), :] = top - bot

    # 10011	Labor	(add all 6000s)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 6000) & 
        (pivot_table.index.get_level_values('Account_Num') < 6999)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(10011, 'Labor'), :] = summed_values

    # 10012	Controllable	(add all 7000s that are CONTROLABLE)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 7000) & 
        (pivot_table.index.get_level_values('Account_Num') < 7999)
    ]
    # add along columns if Controllable
    existing_accounts = filtered_pivot_table.index.get_level_values(1)
    controllable_rows = controlmap[(controlmap['Account'].isin(existing_accounts)) & (controlmap['Controllable'] == 'Controllable')]
    account_names = controllable_rows['Account'].tolist()
    account_names = [(int(x[:4]), x) for x in account_names]
    selected_rows = filtered_pivot_table.loc[account_names,:]
    sum_along_columns = selected_rows.sum()
    pivot_table.loc[(10012, 'Controllable'), :] = sum_along_columns

    # 10013	Uncontrollable	(add all 7000s that are UNCONTROLABLE)
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 7000) & 
        (pivot_table.index.get_level_values('Account_Num') < 7999)
    ]
    # add along columns if Controllable
    existing_accounts = filtered_pivot_table.index.get_level_values(1)
    controllable_rows = controlmap[(controlmap['Account'].isin(existing_accounts)) & (controlmap['Controllable'] == 'Uncontrollable')]
    account_names = controllable_rows['Account'].tolist()
    account_names = [(int(x[:4]), x) for x in account_names]
    selected_rows = filtered_pivot_table.loc[account_names,:]
    sum_along_columns = selected_rows.sum()
    pivot_table.loc[(10013, 'Uncontrollable'), :] = sum_along_columns

    # 10021	Labor %	10011 divide by  4999  (4999, 'Total Income')
    top = pivot_table.loc[(10011, 'Labor'), :]
    bot = pivot_table.loc[(4999, 'Total Income'), :]
    pivot_table.loc[(10021, 'Labor %'), :] = top / bot

    # 10022	Controllable %	10012 divide by  4999  (4999, 'Total Income')
    top = pivot_table.loc[(10012, 'Controllable'), :]
    bot = pivot_table.loc[(4999, 'Total Income'), :]
    pivot_table.loc[(10022, 'Controllable %'), :] = top / bot

    # 10023	Uncontrollable %	10013 divide by  4999  (4999, 'Total Income')
    top = pivot_table.loc[(10013, 'Uncontrollable'), :]
    bot = pivot_table.loc[(4999, 'Total Income'), :]
    pivot_table.loc[(10023, 'Uncontrollable %'), :] = top / bot

    # Sort the index
    pivot_table.sort_index(inplace=True)

    return(pivot_table)


def create_T5_BAL_pivot_table(result_df, myloc):
    '''
    Takes the result dataframe and the T5 location and returns 
    a pivot dataframe across all months
    '''
    # import pandas as pd
    import numpy as np

    result = result_df
    # Pivot the DataFrame by 'Month', keeping columns ('Account_Num', 'Account'), summing 'value', and filtering by 'location'
    pivot_table = result[result['location'] == myloc].pivot_table(index=['Account_Num', 'Account'], columns='monthdt', values='value', aggfunc='sum')

    # print(pivot_table.tail(20))

    ### change the dt to Mmmm YYYY 
    # pivot_table.replace(0.0, np.nan, inplace=True)
    pivot_table = pivot_table.reindex(columns=pivot_table.columns.sort_values())
    pivot_table.columns = pivot_table.columns.strftime('%b %y')

    ### 1199  Total Bank Accounts	sum all 1001 through 1042
    # Filter the pivot table by "Account_Num" greater than 4000 and less than 4999
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1000) & 
        (pivot_table.index.get_level_values('Account_Num') < 1199)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1199, 'Total Bank Accounts'), :] = summed_values

    ### 1498  Total Other Current Assets	sum all 1200 through 1425
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1200) & 
        (pivot_table.index.get_level_values('Account_Num') < 1426)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1498, 'Total Other Current Assets'), :] = summed_values

    ### 1499 Total Current Assets	add 1199 and 1498
    top = pivot_table.loc[(1199, 'Total Bank Accounts'), :]
    bot = pivot_table.loc[(1498, 'Total Other Current Assets'), :]
    pivot_table.loc[(1499, 'Total Current Assets'), :] = top + bot

    ### 1699  Total Fixed Assets	sum 1500 through 1650
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1500) & 
        (pivot_table.index.get_level_values('Account_Num') < 1690)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1699, 'Total Fixed Assets'), :] = summed_values

    ### 1998  Total Other Assets	sum 1700 through 1980
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1700) & 
        (pivot_table.index.get_level_values('Account_Num') < 1981)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1998, 'Total Other Assets'), :] = summed_values

    ### 1999  TOTAL ASSETS	add 1499 + 1699 + 1998
    top = pivot_table.loc[(1499, 'Total Current Assets'), :]
    mid = pivot_table.loc[(1699, 'Total Fixed Assets')]
    bot = pivot_table.loc[(1998, 'Total Other Assets'), :]
    pivot_table.loc[(1999, 'TOTAL ASSETS'), :] = top + mid + bot

    ### 2200  Total Accounts Payable	sum all 2100s
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2100) & 
        (pivot_table.index.get_level_values('Account_Num') < 2200)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2200, 'Total Accounts Payable'), :] = summed_values

    ### 2099 Total Credit Cards	sum all 2000s
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2002) & 
        (pivot_table.index.get_level_values('Account_Num') < 2099)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2099, 'Total Credit Cards'), :] = summed_values


    ### 2798   Total Other Current Liabilities	sum all 2205 through 2730 
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2201) & 
        (pivot_table.index.get_level_values('Account_Num') < 2798)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2798, 'Total Other Current Liabilities'), :] = summed_values

    ### 2799  Total Current Liabilities	add 2200 + 2099 + 2798
    top = pivot_table.loc[(2200, 'Total Accounts Payable'), :]
    mid = pivot_table.loc[(2099, 'Total Credit Cards')]
    bot = pivot_table.loc[(2798, 'Total Other Current Liabilities'), :]
    pivot_table.loc[(2799, 'Total Current Liabilities'), :] = top + mid + bot

    ### 2998  Total Long-Term Liabilities	sum 2910
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2900) & 
        (pivot_table.index.get_level_values('Account_Num') < 2998)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2998, 'Total Long-Term Liabilities'), :] = summed_values

    ### 2999   Total Liabilities	add 2799 + 2998
    top = pivot_table.loc[(2799, 'Total Current Liabilities'), :]
    bot = pivot_table.loc[(2998, 'Total Long-Term Liabilities'), :]
    pivot_table.loc[(2999, 'Total Liabilities'), :] = top + bot

    ### 3998 Total Equity	sum 3000 through 3997
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 3000) & 
        (pivot_table.index.get_level_values('Account_Num') < 3998)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(3998, 'Total Equity'), :] = summed_values

    ### 3999   TOTAL LIABILITIES AND EQUITY	add 3998 + 2999
    top = pivot_table.loc[(2999, 'Total Liabilities'), :]
    bot = pivot_table.loc[(3998, 'Total Equity'), :]
    pivot_table.loc[(3999, 'TOTAL LIABILITIES AND EQUITY'), :] = top + bot

    # remove all 0.0s and replace with NaNs
    pivot_table.replace(0.0, np.nan, inplace=True)

    # Sort the index
    pivot_table.sort_index(inplace=True)

    return(pivot_table)


def create_T5_BAL_OPCO_pivot_table(result_df):
    '''
    Takes the result dataframe and the T5 location and returns 
    a pivot dataframe across all months
    '''
    # import pandas as pd
    import numpy as np

    result = result_df
    # Pivot the DataFrame by 'Month', keeping columns ('Account_Num', 'Account'), summing 'value', and filtering by 'location'
    pivot_table = result.pivot_table(index=['Account_Num', 'Account'], columns='monthdt', values='value', aggfunc='sum')

    # print(pivot_table.tail(20))

    ### change the dt to Mmmm YYYY 
    # pivot_table.replace(0.0, np.nan, inplace=True)
    pivot_table = pivot_table.reindex(columns=pivot_table.columns.sort_values())
    pivot_table.columns = pivot_table.columns.strftime('%b %y')

    ### 1199  Total Bank Accounts	sum all 1001 through 1042
    # Filter the pivot table by "Account_Num" greater than 4000 and less than 4999
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1000) & 
        (pivot_table.index.get_level_values('Account_Num') < 1199)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1199, 'Total Bank Accounts'), :] = summed_values

    ### 1498  Total Other Current Assets	sum all 1200 through 1425
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1200) & 
        (pivot_table.index.get_level_values('Account_Num') < 1426)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1498, 'Total Other Current Assets'), :] = summed_values

    ### 1499 Total Current Assets	add 1199 and 1498
    top = pivot_table.loc[(1199, 'Total Bank Accounts'), :]
    bot = pivot_table.loc[(1498, 'Total Other Current Assets'), :]
    pivot_table.loc[(1499, 'Total Current Assets'), :] = top + bot

    ### 1699  Total Fixed Assets	sum 1500 through 1650
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1500) & 
        (pivot_table.index.get_level_values('Account_Num') < 1690)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1699, 'Total Fixed Assets'), :] = summed_values

    ### 1998  Total Other Assets	sum 1700 through 1980
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 1700) & 
        (pivot_table.index.get_level_values('Account_Num') < 1981)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(1998, 'Total Other Assets'), :] = summed_values

    ### 1999  TOTAL ASSETS	add 1499 + 1699 + 1998
    top = pivot_table.loc[(1499, 'Total Current Assets'), :]
    mid = pivot_table.loc[(1699, 'Total Fixed Assets')]
    bot = pivot_table.loc[(1998, 'Total Other Assets'), :]
    pivot_table.loc[(1999, 'TOTAL ASSETS'), :] = top + mid + bot

    ### 2200  Total Accounts Payable	sum all 2100s
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2100) & 
        (pivot_table.index.get_level_values('Account_Num') < 2200)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2200, 'Total Accounts Payable'), :] = summed_values

    ### 2099 Total Credit Cards	sum all 2000s
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2002) & 
        (pivot_table.index.get_level_values('Account_Num') < 2099)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2099, 'Total Credit Cards'), :] = summed_values


    ### 2798   Total Other Current Liabilities	sum all 2205 through 2730 
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2201) & 
        (pivot_table.index.get_level_values('Account_Num') < 2798)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2798, 'Total Other Current Liabilities'), :] = summed_values

    ### 2799  Total Current Liabilities	add 2200 + 2099 + 2798
    top = pivot_table.loc[(2200, 'Total Accounts Payable'), :]
    mid = pivot_table.loc[(2099, 'Total Credit Cards')]
    bot = pivot_table.loc[(2798, 'Total Other Current Liabilities'), :]
    pivot_table.loc[(2799, 'Total Current Liabilities'), :] = top + mid + bot

    ### 2998  Total Long-Term Liabilities	sum 2910
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 2900) & 
        (pivot_table.index.get_level_values('Account_Num') < 2998)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(2998, 'Total Long-Term Liabilities'), :] = summed_values

    ### 2999   Total Liabilities	add 2799 + 2998
    top = pivot_table.loc[(2799, 'Total Current Liabilities'), :]
    bot = pivot_table.loc[(2998, 'Total Long-Term Liabilities'), :]
    pivot_table.loc[(2999, 'Total Liabilities'), :] = top + bot

    ### 3998 Total Equity	sum 3000 through 3997
    filtered_pivot_table = pivot_table[
        (pivot_table.index.get_level_values('Account_Num') > 3000) & 
        (pivot_table.index.get_level_values('Account_Num') < 3998)
    ]
    summed_values = filtered_pivot_table.sum(axis=0)   # Sum up along the columns
    pivot_table.loc[(3998, 'Total Equity'), :] = summed_values

    ### 3999   TOTAL LIABILITIES AND EQUITY	add 3998 + 2999
    top = pivot_table.loc[(2999, 'Total Liabilities'), :]
    bot = pivot_table.loc[(3998, 'Total Equity'), :]
    pivot_table.loc[(3999, 'TOTAL LIABILITIES AND EQUITY'), :] = top + bot

    # remove all 0.0s and replace with NaNs
    pivot_table.replace(0.0, np.nan, inplace=True)

    # Sort the index
    pivot_table.sort_index(inplace=True)

    return(pivot_table)

